import openpyxl as xl
import os
from pyfirmata import Arduino, util
import time

#Person's name
folder = '/Basma'
path = 'C:/Users/Basma/Downloads/GraduationProject/hearMeOut1'
folder_path = os.path.join(path, folder)

#dedicating a folder for each person
if not os.path.exists(folder_path):
    os.mkdir(folder_path)
os.chdir(folder_path)

#Creating a new file for each data collected iteration
sign = ''                #The word or expression to be signed
t = time.localtime()        #current date and time (to avoid repeated names or overwriting data)
now = time.strftime("%d%m_%H%M%S", t)

wb = xl.Workbook()
ws = wb.active
ws.title = "Sheet1"

#Putting column titles for data
mid = ws.cell(1, 1)
thumb = ws.cell(1, 2)
elbow = ws.cell(1, 3)

mid.value = "mid"
thumb.value = "thumb"
elbow.value = "elbow"

#Connecting with arduino
ard = Arduino('COM3')
print("Communication successfully started")

it = util.Iterator(ard)
it.start()

flex15 = ard.get_pin('a:0:i')   #default15 = 417.2
flex24 = ard.get_pin('a:1:i')   #default24 = 428.5
flex11 = ard.get_pin('a:2:i')   #default11 = 299.5

time.sleep(2)
#Values decrease with time
thresh15 = 390
thresh24 = 390
thresh11 = 280

sensor0 = flex15.read()*500
sensor1 = flex24.read()*500
sensor2 = flex11.read()*500
time.sleep(1)
print(f'flex15: {sensor0}')
print(f'flex24: {sensor1}')
print(f'flex11: {sensor2}')

loop = 'peep'

while (loop != 'quit'):
    print(f'flex15: {flex15.read()*500}')
    print(f'flex24: {flex24.read()*500}')
    print(f'flex11: {flex11.read()*500}')

    # Creating a new file for each data collected iteration
    t = time.localtime()  # current date and time (to avoid repeated names or overwriting data)
    now = time.strftime("%d%m_%H%M%S", t)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Putting column titles for data
    mid = ws.cell(1, 1)
    thumb = ws.cell(1, 2)
    elbow = ws.cell(1, 3)

    mid.value = "mid"
    thumb.value = "thumb"
    elbow.value = "elbow"

    if((flex15.read()*500 < thresh15) or (flex24.read()*500 < thresh24) or (flex11.read()*500 < thresh11)):
        for row in range(2,21):

            mid = ws.cell(row, 1)
            thumb = ws.cell(row, 2)
            elbow = ws.cell(row, 3)

            mid.value = flex15.read() * 500
            thumb.value = flex24.read() * 500
            elbow.value = flex11.read() * 500

            # print(f'flex15: {(ws.cell(2, 1)).value}   flex24: {(ws.cell(2, 2)).value}     flex11: {(ws.cell(2, 3)).value}')
            time.sleep(0.05)
        wb.save(sign + now + '.xlsx')
        loop = input('another iteration? ')


    time.sleep(0.1)