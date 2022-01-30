import openpyxl as xl
from os import listdir
from os.path import isfile, join

mypath = "C:/Users/Basma/Downloads/GraduationProject/hearMeOut1/DataCleaned/Maria"

onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# print(onlyfiles[0])

for exl in onlyfiles:
    file = join(mypath, exl)
    wb = xl.load_workbook(file)
    sheet = wb['Sheet1']


    for row in range(2, sheet.max_row+1):
        # print(sheet.cell(row,3).value)
        if sheet.cell(row,1).value <= 300:
            print(f'{file}')
            sheet.cell(row, 1).value = (sheet.cell(row-1,1).value+ sheet.cell(row+1,1).value)/2


        if sheet.cell(row,2).value <= 300:
            print(f'{file}')
            sheet.cell(row,2).value = (sheet.cell(row-1,2).value + sheet.cell(row + 1, 2).value) / 2


        if sheet.cell(row,3).value <= 150:
            print(f'{file}')
            sheet.cell(row,3).value = (sheet.cell(row-1,3).value + sheet.cell(row + 1, 3).value) / 2

    wb.template = False
    wb.save(file)
# print(f'{cell1},\n {cell2},\n {cell3}')
